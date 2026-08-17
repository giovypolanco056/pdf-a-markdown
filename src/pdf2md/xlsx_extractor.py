"""Lector de Excel (.xlsx) -> DocumentResult.

Cada hoja del libro se convierte en una tabla Markdown (precedida de un
encabezado con el nombre de la hoja cuando hay más de una). Produce los mismos
``Block`` que los demás lectores, así que el resto del pipeline se reutiliza.
Se apoya en ``openpyxl``.
"""
from __future__ import annotations

import datetime as _dt
from pathlib import Path

from .config import Config
from .errors import PDFConversionError
from .models import Block, BlockType, DocumentResult, Page

try:
    from openpyxl import load_workbook
except ImportError:  # pragma: no cover
    load_workbook = None

_MAX_ROWS = 5000       # límite de seguridad por hoja (evita Markdown gigantes)
_GENERIC = {"sheet", "sheet1", "hoja", "hoja1", "libro1"}


def _require_openpyxl() -> None:
    if load_workbook is None:
        raise PDFConversionError(
            "Falta la librería 'openpyxl'. Instálala con:  pip install openpyxl")


def _fmt(value) -> str:
    if value is None:
        return ""
    if isinstance(value, bool):
        return "Sí" if value else "No"
    if isinstance(value, float):
        return str(int(value)) if value.is_integer() else f"{value:g}"
    if isinstance(value, (_dt.datetime, _dt.date)):
        return value.isoformat().replace("T00:00:00", "")
    return str(value).strip()


def _sheet_rows(ws) -> tuple[list[list[str]], bool]:
    raw: list[list[str]] = []
    truncated = False
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i >= _MAX_ROWS:
            truncated = True
            break
        raw.append([_fmt(v) for v in row])

    # Recortar columnas vacías al final (openpyxl suele inflar las dimensiones).
    last_col = 0
    for r in raw:
        for j in range(len(r) - 1, -1, -1):
            if r[j] != "":
                last_col = max(last_col, j + 1)
                break
    raw = [r[:last_col] for r in raw]

    # Quitar filas totalmente vacías al principio y al final.
    while raw and all(c == "" for c in raw[-1]):
        raw.pop()
    while raw and all(c == "" for c in raw[0]):
        raw.pop(0)
    return raw, truncated


def convert_xlsx(path: Path, config: Config) -> DocumentResult:
    _require_openpyxl()
    path = Path(path)
    try:
        wb = load_workbook(str(path), data_only=True, read_only=True)
    except Exception as exc:
        raise PDFConversionError(f"No se pudo abrir el Excel: {exc}") from exc

    result = DocumentResult(source_path=str(path))
    page = Page(number=1, source="xlsx")
    blocks: list[Block] = []
    warnings: list[str] = []
    try:
        sheets = wb.worksheets
        multi = len(sheets) > 1
        for ws in sheets:
            rows, truncated = _sheet_rows(ws)
            if not rows:
                continue
            name = str(ws.title or "").strip()
            if multi or name.lower() not in _GENERIC:
                blocks.append(Block(type=BlockType.HEADING, text=name or "Hoja", level=2))
            blocks.append(Block(type=BlockType.TABLE, rows=rows))
            if truncated:
                warnings.append(
                    f"La hoja '{name}' supera {_MAX_ROWS} filas; se truncó el resto.")
        sheet_count = len(sheets)
    finally:
        wb.close()

    if not blocks:
        blocks.append(Block(type=BlockType.PARAGRAPH,
                            text="> ⚠️ El Excel no contenía datos legibles."))
        warnings.append("El libro no tenía celdas con datos.")

    page.blocks = blocks
    page.warnings = warnings
    result.pages.append(page)
    result.warnings = list(warnings)
    result.metadata = {
        "title": path.stem,
        "source": path.name,
        "file_type": "xlsx",
        "sheets": sheet_count,
        "ocr": False,
        "language": config.language,
        "date_processed": _dt.date.today().isoformat(),
        "tags": [],
    }
    return result
